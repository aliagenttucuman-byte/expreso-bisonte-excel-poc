"""
Merger INICIAL + SISTEMA = FINAL — Expreso Bisonte Cobranzas Contado
=====================================================================
Reglas de negocio (relevadas con Edith, gerenta Bisonte — jun 2026):

INICIAL  = planilla acumulada que viene trabajando Edith (con anotaciones manuales)
SISTEMA  = descarga fresca de Transoft del dia de hoy
FINAL    = merge inteligente de ambas

REGLAS:
  1. Existentes (en INICIAL y en SISTEMA): van a FINAL con anotaciones de INICIAL preservadas.
     Si el estado cambio en SISTEMA respecto a INICIAL → pintar fila en ROJO (alerta Edith).
  2. Nuevos (solo en SISTEMA): van a FINAL sin anotaciones (celdas vacías) → carga manual.
     Se pintan en AMARILLO para que el equipo los identifique fácil.
  3. Eliminados (en INICIAL pero ya NO en SISTEMA): se descartan (ya cobrados/cerrados).
  4. Columna nueva "DIAS_ATRASO": fecha hoy − fechaedit (última edición en Transoft).
  5. SISTEMA se filtra por estado=ED antes del merge (otras matrices manejan el resto).
  6. Duplicados en SISTEMA se eliminan — se queda con la primera ocurrencia por nro.

CAMPOS MANUALES (solo en INICIAL/FINAL, no en SISTEMA):
  JUSTIFICACIÓN, REFERENTE, ESTADO, OBSERVACIÓN

COLORES:
  ROJO   (#FF0000) — estado cambió en Transoft, O días de atraso superan tolerancia
                     (CC: >4 días, resto: >7 días)
  AMARILLO (#FFFF00) — guía nueva del día → cargar manualmente
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colores ──────────────────────────────────────────────────────────────────
FILL_ROJO     = PatternFill("solid", fgColor="FFFF0000")   # estado cambió
FILL_AMARILLO = PatternFill("solid", fgColor="FFFFFF00")   # nuevo del día
FILL_HEADER   = PatternFill("solid", fgColor="FF2E4A8C")   # encabezado azul Bisonte
FONT_HEADER   = Font(bold=True, color="FFFFFFFF", size=10)
FONT_NORMAL   = Font(size=9)
ALIGN_CENTER  = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center")

# ── Columnas esperadas ────────────────────────────────────────────────────────
# SISTEMA tiene estas columnas (sin las manuales)
SISTEMA_COLS = [
    "nro", "guiafec", "razsocc", "clase", "fechaedit",
    "sucori", "sucdest", "importe", "saldo", "succobro",
    "estado", "tiporec", "sucursal", "nrogen_a",
]

# Columnas manuales que agrega Edith (preservar de INICIAL)
MANUAL_COLS = ["JUSTIFICACIÓN", "REFERENTE", "ESTADO", "OBSERVACIÓN"]

# Orden final de columnas en FINAL
FINAL_COLS = [
    "nro",
    "JUSTIFICACIÓN",
    "REFERENTE",
    "ESTADO",
    "OBSERVACIÓN",
    "DIAS_ATRASO",
    "guiafec",
    "razsocc",
    "clase",
    "fechaedit",
    "sucori",
    "sucdest",
    "importe",
    "saldo",
    "succobro",
    "tiporec",
    "sucursal",
    "nrogen_a",
    "__ORIGEN__",   # NUEVO | EXISTENTE | EXISTENTE_CAMBIO — para badge en UI
]

# Índice (1-based) de fechaedit en FINAL_COLS — para pintar la celda por columna
_FECHAEDIT_COL_IDX = FINAL_COLS.index("fechaedit") + 1 if "fechaedit" in FINAL_COLS else None

# ── Helpers ───────────────────────────────────────────────────────────────────

def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, key_col: str = "nro") -> int:
    """Detecta la fila de encabezados buscando 'nro' en las primeras 5 filas."""
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip().lower() == key_col.lower():
                return r
    return 1


def _sheet_to_dicts(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[str], list[dict]]:
    """Convierte una sheet en lista de dicts, detectando la fila de headers automáticamente."""
    header_row = _find_header_row(ws)
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        headers.append(str(v).strip() if v is not None else f"__col{c}__")

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row: dict[str, Any] = {}
        for c, h in enumerate(headers, 1):
            row[h] = ws.cell(r, c).value
        # Solo incluir filas con nro válido (empieza con A., B. o R.)
        nro = row.get("nro") or row.get("NRO") or ""
        if nro and str(nro).strip()[:2] in ("A.", "B.", "R."):
            row["nro"] = str(nro).strip()
            rows.append(row)
    return headers, rows


def _calc_dias_atraso(fechaedit: Any) -> int | str:
    """Calcula días desde fechaedit hasta hoy (hora Argentina). Retorna int o '' si no hay fecha."""
    if not fechaedit:
        return ""
    try:
        if isinstance(fechaedit, datetime):
            fecha = fechaedit.date()
        elif isinstance(fechaedit, date):
            fecha = fechaedit
        else:
            s = str(fechaedit).strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    fecha = datetime.strptime(s[:10], fmt).date()
                    break
                except ValueError:
                    continue
            else:
                return ""
        # Fecha de hoy en Argentina (UTC-3)
        from datetime import timezone, timedelta
        ar_tz = timezone(timedelta(hours=-3))
        hoy = datetime.now(ar_tz).date()
        return (hoy - fecha).days
    except Exception:
        return ""


def _es_rojo_por_atraso(dias_atraso: Any, succobro: Any) -> bool:
    """
    Regla de Edith — tolerancia por sucursal:
      - Casa Central (CC): cualquier día de atraso → rojo (tolerancia = 0)
      - Resto de sucursales: más de 7 días → rojo
    """
    if not isinstance(dias_atraso, int):
        return False
    suc = str(succobro or "").strip().upper()
    tolerancia = 0 if suc == "CC" else 7
    return dias_atraso > tolerancia


def _normalize_estado(v: Any) -> str:
    """Normaliza el valor del estado para comparar."""
    if v is None:
        return ""
    return str(v).strip().upper()


def _hay_diferencia(importe: Any, saldo: Any) -> bool:
    """Detecta si importe != saldo (diferencia real, no redondeo de centavos)."""
    try:
        imp = float(importe) if importe not in (None, "", "nan") else None
        sal = float(saldo)   if saldo   not in (None, "", "nan") else None
        if imp is None or sal is None:
            return False
        return abs(imp - sal) > 0.01  # tolerancia 1 centavo
    except (TypeError, ValueError):
        return False


# ── Función principal ─────────────────────────────────────────────────────────

def merge_contado(
    inicial_bytes: bytes,
    sistema_bytes: bytes,
    sheet_inicial: str = "INICIAL",
    sheet_sistema: str = "SISTEMA",
) -> bytes:
    """
    Recibe los bytes de INICIAL y SISTEMA (Excel), genera FINAL y retorna sus bytes.

    Args:
        inicial_bytes: contenido del xlsx INICIAL
        sistema_bytes: contenido del xlsx SISTEMA
        sheet_inicial: nombre de la hoja en el archivo INICIAL
        sheet_sistema: nombre de la hoja en el archivo SISTEMA

    Returns:
        bytes del xlsx FINAL generado
    """
    # Cargar workbooks
    wb_ini = load_workbook(io.BytesIO(inicial_bytes), data_only=True)
    wb_sis = load_workbook(io.BytesIO(sistema_bytes), data_only=True)

    # Detectar sheets — buscar case-insensitive
    def get_sheet(wb: Workbook, name: str):
        for sn in wb.sheetnames:
            if sn.strip().upper() == name.strip().upper():
                return wb[sn]
        # Si no encuentra, usar la primera
        return wb.active

    ws_ini = get_sheet(wb_ini, sheet_inicial)
    ws_sis = get_sheet(wb_sis, sheet_sistema)

    _, rows_ini = _sheet_to_dicts(ws_ini)
    _, rows_sis_raw = _sheet_to_dicts(ws_sis)

    # Sin filtro de estado — traer todos los registros del SISTEMA
    # (el foco operativo es ED, pero Edith necesita ver el panorama completo)

    # Deduplicar SISTEMA por nro (pago a cuenta + nota de crédito generan duplicados)
    # Se queda con la primera ocurrencia
    rows_sis: list[dict] = []
    seen_nros: set[str] = set()
    for r in rows_sis_raw:
        if r["nro"] not in seen_nros:
            seen_nros.add(r["nro"])
            rows_sis.append(r)

    # Indexar por nro
    dict_ini: dict[str, dict] = {r["nro"]: r for r in rows_ini}
    dict_sis: dict[str, dict] = {r["nro"]: r for r in rows_sis}

    nros_ini = set(dict_ini.keys())
    nros_sis = set(dict_sis.keys())

    # Clasificar
    nros_existentes = nros_ini & nros_sis          # preservar anotaciones
    nros_nuevos     = nros_sis - nros_ini          # nuevos del día → amarillo
    # nros_eliminados = nros_ini - nros_sis        # cobrados → descartar

    # Estadísticas para el resumen
    stats = {
        "existentes": len(nros_existentes),
        "nuevos":     len(nros_nuevos),
        "eliminados": len(nros_ini - nros_sis),
        "estado_cambio": 0,
    }

    # ── Construir FINAL ───────────────────────────────────────────────────────
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "FINAL"

    # Escribir encabezado
    ws_out.row_dimensions[1].height = 20
    for c, col_name in enumerate(FINAL_COLS, 1):
        cell = ws_out.cell(1, c, col_name)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    # (sin columnas extra)

    row_out = 2

    # Helper para escribir una fila
    def write_row(row_data: dict, fill: PatternFill | None = None):
        nonlocal row_out
        for c, col_name in enumerate(FINAL_COLS, 1):
            v = row_data.get(col_name, "")
            # Limpiar #N/A
            if str(v) in ("#N/A", "#NA", "nan"):
                v = ""
            # Redondear importe y saldo a 2 decimales
            if col_name in ("importe", "saldo") and v not in (None, ""):
                try:
                    v = round(float(v), 2)
                except (TypeError, ValueError):
                    pass
            cell = ws_out.cell(row_out, c, v)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_LEFT
            if col_name in ("importe", "saldo") and isinstance(v, float):
                cell.number_format = '#,##0.00'
            if fill:
                cell.fill = fill

        # Pintar celda fechaedit de rojo si el atraso supera tolerancia por sucdest
        # CC → > 2 días, resto → > 7 días
        dias    = row_data.get("DIAS_ATRASO", "")
        sucdest = str(row_data.get("sucdest", "") or "").strip().upper()
        if isinstance(dias, int) and _FECHAEDIT_COL_IDX:
            tolerancia = 0 if sucdest == "CC" else 7
            if dias > tolerancia:
                ws_out.cell(row_out, _FECHAEDIT_COL_IDX).fill = FILL_ROJO

        # (sin columnas extra)

        row_out += 1

    # 1. EXISTENTES — preservar anotaciones, detectar cambio de estado
    for nro in sorted(nros_existentes):
        r_ini = dict_ini[nro]
        r_sis = dict_sis[nro]

        estado_ini = _normalize_estado(r_ini.get("ESTADO") or r_ini.get("estado"))
        estado_sis = _normalize_estado(r_sis.get("estado") or r_sis.get("ESTADO"))
        cambio = estado_ini and estado_sis and estado_ini != estado_sis

        if cambio:
            stats["estado_cambio"] += 1

        row_data: dict[str, Any] = {}

        # Datos del sistema (frescos)
        for col in SISTEMA_COLS:
            row_data[col] = r_sis.get(col, "")

        # Anotaciones manuales de INICIAL (preservar)
        for col in MANUAL_COLS:
            v = r_ini.get(col, "")
            if str(v) in ("#N/A", "#NA", "nan", "None"):
                v = ""
            row_data[col] = v

        # ESTADO siempre desde SISTEMA (lo pone Transoft, no Edith)
        row_data["ESTADO"] = estado_sis

        # REFERENTE: si viene vacío del INICIAL → auto-sugerir desde succobro
        if not str(row_data.get("REFERENTE", "") or "").strip():
            succobro_ref = str(r_sis.get("succobro", "") or "").strip().upper()
            if succobro_ref:
                row_data["REFERENTE"] = succobro_ref
                stats["referente_auto"] = stats.get("referente_auto", 0) + 1

        # Auto-detectar VER DIF: si importe != saldo y OBSERVACIÓN vacía → sugerir
        importe = r_sis.get("importe")
        saldo   = r_sis.get("saldo")
        obs_actual = str(row_data.get("OBSERVACIÓN", "") or "").strip()
        if _hay_diferencia(importe, saldo) and not obs_actual:
            row_data["OBSERVACIÓN"] = "VER DIF"
            stats["ver_dif_auto"] = stats.get("ver_dif_auto", 0) + 1

        # Días de atraso — desde fechaedit (última edición en Transoft = fecha de entrega)
        row_data["DIAS_ATRASO"] = _calc_dias_atraso(r_sis.get("fechaedit"))

        # Metadatos para columnas extra
        row_data["__estado_sis__"] = estado_sis
        row_data["__origen__"] = "EXISTENTE_CAMBIO" if cambio else "EXISTENTE"
        row_data["__ORIGEN__"] = row_data["__origen__"]

        fill_fila = FILL_ROJO if cambio else None
        write_row(row_data, fill_fila)

    # 2. NUEVOS — solo datos del sistema, campos manuales vacíos
    for nro in sorted(nros_nuevos):
        r_sis = dict_sis[nro]

        row_data = {}
        for col in SISTEMA_COLS:
            row_data[col] = r_sis.get(col, "")

        for col in MANUAL_COLS:
            row_data[col] = ""  # vacío → carga manual

        # Pre-poblar ESTADO con el estado de Transoft para los nuevos
        # Edith lo confirma o corrige — no arranca de cero
        estado_transoft = _normalize_estado(r_sis.get("estado"))
        if estado_transoft:
            row_data["ESTADO"] = estado_transoft
            stats["estado_auto"] = stats.get("estado_auto", 0) + 1

        # Auto-sugerir REFERENTE desde succobro (regla Edith: 90% correlación directa)
        succobro = str(r_sis.get("succobro", "") or "").strip().upper()
        if succobro:
            row_data["REFERENTE"] = succobro
            stats["referente_auto"] = stats.get("referente_auto", 0) + 1

        # Auto-detectar VER DIF en nuevos también
        importe = r_sis.get("importe")
        saldo   = r_sis.get("saldo")
        if _hay_diferencia(importe, saldo):
            row_data["OBSERVACIÓN"] = "VER DIF"
            stats["ver_dif_auto"] = stats.get("ver_dif_auto", 0) + 1

        row_data["DIAS_ATRASO"] = _calc_dias_atraso(r_sis.get("fechaedit"))
        row_data["__estado_sis__"] = _normalize_estado(r_sis.get("estado"))
        row_data["__origen__"] = "NUEVO"
        row_data["__ORIGEN__"] = "NUEVO"

        write_row(row_data, FILL_AMARILLO)

    # ── Ajustar ancho de columnas ─────────────────────────────────────────────
    col_widths = {
        "nro": 22, "JUSTIFICACIÓN": 18, "REFERENTE": 12, "ESTADO": 10,
        "OBSERVACIÓN": 20, "DIAS_ATRASO": 12, "guiafec": 14, "razsocc": 32,
        "clase": 22, "fechaedit": 18, "sucori": 9, "sucdest": 9,
        "importe": 14, "saldo": 14, "succobro": 10, "tiporec": 9,
        "sucursal": 9, "nrogen_a": 12,
    }
    all_cols = FINAL_COLS
    for c, col_name in enumerate(all_cols, 1):
        ws_out.column_dimensions[get_column_letter(c)].width = col_widths.get(col_name, 14)

    # Freeze primera fila
    ws_out.freeze_panes = "A2"

    # ── Hoja de resumen ───────────────────────────────────────────────────────
    ws_res = wb_out.create_sheet("RESUMEN")
    resumen_data = [
        ("RESUMEN DEL MERGE", ""),
        ("Fecha de procesamiento", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("Registros en INICIAL", len(rows_ini)),
        ("Registros en SISTEMA", len(rows_sis)),
        ("Registros en FINAL",   stats["existentes"] + stats["nuevos"]),
        ("", ""),
        ("Existentes (anotaciones preservadas)", stats["existentes"]),
        ("  → Con cambio de estado (ROJO)", stats["estado_cambio"]),
        ("  → Sin cambio de estado", stats["existentes"] - stats["estado_cambio"]),
        ("Nuevos del día (AMARILLO — carga manual)", stats["nuevos"]),
        ("Eliminados (cobrados/cerrados)", stats["eliminados"]),
        ("", ""),
        ("LEYENDA DE COLORES", ""),
        ("ROJO", "Estado cambió en Transoft, O atraso > 4 días (CC) / > 7 días (otras sucursales)"),
        ("AMARILLO", "Guía nueva del día — completar manualmente (dentro de tolerancia)"),
        ("Sin color", "Registro existente sin cambios"),
    ]
    for i, (k, v) in enumerate(resumen_data, 1):
        ws_res.cell(i, 1, k).font = Font(bold=(i == 1 or k.isupper()), size=10)
        ws_res.cell(i, 2, v).font = Font(size=10)
    ws_res.column_dimensions["A"].width = 45
    ws_res.column_dimensions["B"].width = 45

    # Guardar y retornar bytes
    buf = io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf.read(), stats


# ─────────────────────────────────────────────────────────────────────────────
# excel_to_table — parsea un Excel FINAL para la tabla editable del frontend
# ─────────────────────────────────────────────────────────────────────────────

def excel_to_table(data: bytes, sheet_name: str = "FINAL") -> dict:
    """
    Lee un Excel y devuelve:
      { columns: [...], rows: [{_row_idx, _color, col1: val, col2: val, ...}] }
    _color: 'red' | 'yellow' | 'none'  (basado en el fill de la fila)
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)

    # Buscar la hoja FINAL, o la primera disponible
    ws = None
    for name in [sheet_name, "FINAL", "CONTADO", wb.sheetnames[0]]:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        raise ValueError(f"No se encontró hoja '{sheet_name}' en el Excel")

    # Leer encabezados — buscar la fila que tenga al menos 5 celdas con valor
    header_row = None
    for row in ws.iter_rows():
        vals = [c.value for c in row if c.value is not None]
        if len(vals) >= 5:
            header_row = row
            break
    if header_row is None:
        return {"columns": [], "rows": []}

    columns = [str(c.value).strip() if c.value is not None else f"COL_{c.column}" for c in header_row]
    header_row_num = header_row[0].row

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_num + 1)):
        vals = [c.value for c in row]
        if not any(v is not None for v in vals):
            continue

        # Detectar color de la fila (primera celda con fill)
        color = "none"
        for cell in row:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.type == "rgb":
                rgb = fill.fgColor.rgb  # e.g. "FFFF0000" o "FFFFFF00"
                if rgb in ("FFFF0000", "FF FF0000", "FFFF0000".lower()):
                    color = "red"
                elif rgb in ("FFFFFF00", "FFFFFF00".lower(), "FFFFFF0000"):
                    color = "yellow"
                elif rgb.upper().endswith("FF0000"):
                    color = "red"
                elif rgb.upper().endswith("FFFF00"):
                    color = "yellow"
                if color != "none":
                    break

        row_dict: dict = {"_row_idx": i, "_color": color}
        for col_name, cell in zip(columns, row):
            val = cell.value
            if isinstance(val, (datetime, date)):
                val = val.strftime("%Y-%m-%d")
            elif val is None:
                val = ""
            else:
                val = str(val)
            row_dict[col_name] = val

        rows.append(row_dict)

    return {"columns": columns, "rows": rows}


# ─────────────────────────────────────────────────────────────────────────────
# table_to_excel — convierte las ediciones del frontend a Excel descargable
# ─────────────────────────────────────────────────────────────────────────────

def table_to_excel(columns: list[str], rows: list[dict], orig_bytes: bytes | None = None) -> bytes:
    """
    Genera un Excel a partir de columns + rows editados.
    Si orig_bytes está presente, preserva el estilo del Excel original
    reemplazando solo los valores de las celdas de datos.
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from copy import copy

    RED_FILL    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FF1E293B", end_color="FF1E293B", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=10)
    DATA_FONT   = Font(size=10)

    wb = Workbook()
    ws = wb.active
    ws.title = "FINAL"

    # Encabezados
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[cell.column_letter].width = max(12, len(col_name) + 4)

    # Índices de columnas con reglas de color (1-based)
    def col_idx(name):
        return columns.index(name) + 1 if name in columns else None

    dias_col_idx     = col_idx("DIAS_ATRASO")
    observ_col_idx   = col_idx("OBSERVACIÓN")
    fechaedit_col_idx = col_idx("fechaedit")
    succobro_col_idx  = col_idx("succobro")

    FILL_VERDE   = PatternFill(start_color="FFDCFCE7", end_color="FFDCFCE7", fill_type="solid")
    FILL_AMARILLO_CELDA = PatternFill(start_color="FFFEF9C3", end_color="FFFEF9C3", fill_type="solid")

    # Datos
    for row_idx, row in enumerate(rows, 2):
        for col_idx_iter, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx_iter, value=val if val != "" else None)
            cell.font = DATA_FONT

        # Regla 1 — DIAS_ATRASO: verde / amarillo / rojo según tolerancia
        if dias_col_idx:
            dias_val = row.get("DIAS_ATRASO", "")
            sucdest = str(row.get("sucdest", "") or "").strip().upper()
            tolerancia = 0 if sucdest == "CC" else 7
            try:
                dias_int = int(dias_val)
                celda = ws.cell(row=row_idx, column=dias_col_idx)
                if dias_int > tolerancia:
                    celda.fill = RED_FILL
                elif dias_int > int(tolerancia * 0.7):
                    celda.fill = FILL_AMARILLO_CELDA
                elif dias_int >= 0:
                    celda.fill = FILL_VERDE
            except (ValueError, TypeError):
                pass

        # Regla 2 — OBSERVACIÓN: rojo si valor es "VER DIF"
        if observ_col_idx:
            obs_val = str(row.get("OBSERVACIÓN", "") or "").strip().upper()
            if obs_val == "VER DIF":
                ws.cell(row=row_idx, column=observ_col_idx).fill = RED_FILL

        # Regla 3 — fechaedit: rojo si DIAS_ATRASO supera tolerancia
        if fechaedit_col_idx and dias_col_idx:
            dias_val = row.get("DIAS_ATRASO", "")
            sucdest = str(row.get("sucdest", "") or "").strip().upper()
            tolerancia = 0 if sucdest == "CC" else 7
            try:
                if int(dias_val) > tolerancia:
                    ws.cell(row=row_idx, column=fechaedit_col_idx).fill = RED_FILL
            except (ValueError, TypeError):
                pass

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

